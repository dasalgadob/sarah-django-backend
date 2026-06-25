"""Parses an uploaded .xlsx workbook and upserts ItemPrice records for a company.

Rows are matched by their own 'id' column: blank id creates a new price,
a filled id updates that exact existing record (must belong to the company).
The related Item is resolved by id and must also belong to the company.

IVA, excise tax and total are always recomputed from Precio x the resolved
item's tax rates -- whatever is typed in those columns is ignored on import.
"""

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import openpyxl

from django.db import transaction

from ..models import Item, ItemPrice
from .item_price_excel_columns import (
    COLUMN_ORDER, ID, ITEM_ID, ITEM_PRICE_TYPE, ITEM_PRICES_SHEET_NAME,
    PRICE, PRICE_TYPE_VALUES_BY_LABEL,
)
from .tax_calculations import calculate_taxes

REQUIRED_FIELD_LABEL = 'Este campo es obligatorio.'


class InvalidWorkbookError(Exception):
    """Raised when the uploaded file can't be parsed as an .xlsx workbook."""


class RowValidationError(Exception):
    """Raised for a single row that fails validation; carries the field errors."""

    def __init__(self, errors):
        self.errors = errors
        super().__init__(str(errors))


@dataclass
class RowError:
    row: int
    errors: dict


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list = field(default_factory=list)

    def as_dict(self):
        return {
            'created': self.created,
            'updated': self.updated,
            'errors': [{'row': row_error.row, 'errors': row_error.errors} for row_error in self.errors],
        }


class ItemPriceExcelImporter:
    """Upserts ItemPrice rows parsed from an .xlsx file into a single company."""

    def __init__(self, company_id):
        self.company_id = company_id

    def import_workbook(self, file_obj):
        try:
            workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
            sheet = (
                workbook[ITEM_PRICES_SHEET_NAME]
                if ITEM_PRICES_SHEET_NAME in workbook.sheetnames
                else workbook.active
            )
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
        except Exception as exc:
            raise InvalidWorkbookError(f'No se pudo leer el archivo: {exc}') from exc

        result = ImportResult()
        for row_number, raw_row in enumerate(rows, start=2):
            if raw_row is None or all(cell in (None, '') for cell in raw_row):
                continue
            row_data = dict(zip(COLUMN_ORDER, raw_row))
            try:
                with transaction.atomic():
                    created = self._upsert_row(row_data)
                result.created += int(created)
                result.updated += int(not created)
            except RowValidationError as exc:
                result.errors.append(RowError(row=row_number, errors=exc.errors))

        return result

    def _upsert_row(self, row_data):
        errors = {}

        existing_price = self._resolve_existing_price(row_data.get(ID), errors)
        item = self._resolve_item(row_data.get(ITEM_ID), errors)
        item_price_type = self._resolve_price_type(row_data.get(ITEM_PRICE_TYPE), errors)
        price = self._resolve_price(row_data.get(PRICE), errors)

        if errors:
            raise RowValidationError(errors)

        iva, excise_tax, total = calculate_taxes(item, price)

        if existing_price is not None:
            existing_price.item = item
            existing_price.item_price_type = item_price_type
            existing_price.price = price
            existing_price.iva = iva
            existing_price.excise_tax = excise_tax
            existing_price.total = total
            existing_price.save()
            return False

        ItemPrice.objects.create(
            company_id=self.company_id,
            item=item,
            item_price_type=item_price_type,
            price=price,
            iva=iva,
            excise_tax=excise_tax,
            total=total,
        )
        return True

    def _resolve_existing_price(self, raw_id, errors):
        if raw_id in (None, ''):
            return None
        price_id = self._parse_int(raw_id)
        if price_id is None:
            errors[ID] = f'"{raw_id}" no es un id válido.'
            return None
        price = ItemPrice.objects.filter(pk=price_id, company_id=self.company_id).first()
        if price is None:
            errors[ID] = f'El precio con id {price_id} no existe o no pertenece a esta empresa.'
        return price

    def _resolve_item(self, raw_item_id, errors):
        if raw_item_id in (None, ''):
            errors[ITEM_ID] = REQUIRED_FIELD_LABEL
            return None
        item_id = self._parse_int(raw_item_id)
        if item_id is None:
            errors[ITEM_ID] = f'"{raw_item_id}" no es un id válido.'
            return None
        item = Item.objects.filter(pk=item_id, company_id=self.company_id).first()
        if item is None:
            errors[ITEM_ID] = f'El producto con id {item_id} no existe o no pertenece a esta empresa.'
        return item

    @staticmethod
    def _resolve_price_type(raw_value, errors):
        value = str(raw_value).strip().lower() if raw_value not in (None, '') else ''
        if not value:
            errors[ITEM_PRICE_TYPE] = REQUIRED_FIELD_LABEL
            return None
        price_type = PRICE_TYPE_VALUES_BY_LABEL.get(value)
        if price_type is None:
            errors[ITEM_PRICE_TYPE] = f'"{raw_value}" no es válido. Use "Venta" o "Compra".'
        return price_type

    @staticmethod
    def _resolve_price(raw_value, errors):
        if raw_value in (None, ''):
            errors[PRICE] = REQUIRED_FIELD_LABEL
            return None
        try:
            return Decimal(str(raw_value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            errors[PRICE] = f'"{raw_value}" no es un número válido.'
            return None

    @staticmethod
    def _parse_int(raw_value):
        try:
            return int(raw_value)
        except (TypeError, ValueError):
            return None
