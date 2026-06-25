"""Renders Item querysets as .xlsx workbooks for download.

The workbook also includes reference sheets listing the valid values for
each FK column (group names, unit abbreviations, tax type/rate pairs), so
the same file can be edited and re-uploaded via ItemExcelImporter.
"""

from decimal import Decimal

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from ..models import ItemGroup
from reference_tables.models import ExciseTaxRate, ExciseTaxType, IvaRate, IvaType, UnitMeasure
from .item_excel_columns import HEADER_ROW, ITEMS_SHEET_NAME

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class ItemExcelExporter:
    """Builds an .xlsx workbook from a queryset of Item records."""

    def __init__(self, queryset):
        self.queryset = queryset.select_related(
            'item_group', 'unit_measure', 'iva_type', 'iva_rate',
            'excise_tax_type', 'excise_tax_rate',
        )

    def build_workbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = ITEMS_SHEET_NAME
        sheet.append(HEADER_ROW)

        for item in self.queryset:
            sheet.append(self._row_for(item))

        self._autosize_columns(sheet, HEADER_ROW)
        self._append_reference_sheets(workbook)
        return workbook

    @staticmethod
    def _row_for(item):
        return [
            item.base_code,
            item.base_name,
            item.name,
            item.item_group.name if item.item_group_id else None,
            item.unit_measure.abbreviation if item.unit_measure_id else None,
            item.iva_type.name if item.iva_type_id else None,
            float(item.iva_rate.value) if item.iva_rate_id else None,
            item.excise_tax_type.name if item.excise_tax_type_id else None,
            float(item.excise_tax_rate.value) if item.excise_tax_rate_id else None,
        ]

    def _append_reference_sheets(self, workbook):
        self._add_reference_sheet(
            workbook, 'Grupos', ['Grupo / Categoría'],
            ItemGroup.objects.order_by('name').values_list('name'),
        )
        self._add_reference_sheet(
            workbook, 'Unidades de medida', ['Unidad de medida', 'Nombre'],
            UnitMeasure.objects.order_by('abbreviation').values_list('abbreviation', 'name'),
        )
        self._add_reference_sheet(
            workbook, 'Tipos de IVA', ['Tipo de IVA'],
            IvaType.objects.order_by('name').values_list('name'),
        )
        self._add_reference_sheet(
            workbook, 'Tarifas de IVA', ['Tipo de IVA', 'Tarifa de IVA'],
            IvaRate.objects.select_related('iva_type')
                .order_by('iva_type__name', 'value')
                .values_list('iva_type__name', 'value'),
        )
        self._add_reference_sheet(
            workbook, 'Tipos de imp. al consumo', ['Tipo de impuesto al consumo'],
            ExciseTaxType.objects.order_by('name').values_list('name'),
        )
        self._add_reference_sheet(
            workbook, 'Tarifas de imp. al consumo', ['Tipo de impuesto al consumo', 'Tarifa de impuesto al consumo'],
            ExciseTaxRate.objects.select_related('excise_tax_type')
                .order_by('excise_tax_type__name', 'value')
                .values_list('excise_tax_type__name', 'value'),
        )

    @classmethod
    def _add_reference_sheet(cls, workbook, title, headers, rows):
        sheet = workbook.create_sheet(title=title)
        sheet.append(headers)
        for row in rows:
            sheet.append([float(value) if isinstance(value, Decimal) else value for value in row])
        cls._autosize_columns(sheet, headers)

    @staticmethod
    def _autosize_columns(sheet, headers):
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(len(header) + 2, 18)

    def as_http_response(self, filename='items.xlsx'):
        workbook = self.build_workbook()
        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
