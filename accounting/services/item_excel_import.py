"""Parses an uploaded .xlsx workbook and upserts Item records for a company.

Rows are matched/created by (company, code), where code equals base_code for
the non-variant "parent" item — this bulk tool does not manage item variants.
"""

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import openpyxl

from django.db import transaction

from ..models import Item, ItemGroup
from reference_tables.models import ExciseTaxRate, ExciseTaxType, IvaRate, IvaType, UnitMeasure
from .item_excel_columns import (
    BASE_CODE, BASE_NAME, COLUMN_ORDER, EXCISE_TAX_RATE, EXCISE_TAX_TYPE,
    IVA_RATE, IVA_TYPE, ITEM_GROUP, ITEMS_SHEET_NAME, NAME, UNIT_MEASURE,
)

REQUIRED_FIELD_LABEL = 'Este campo es obligatorio.'


class InvalidWorkbookError(Exception):
    """Raised when the uploaded file can't be parsed as an .xlsx workbook."""


class RowValidationError(Exception):
    """Raised for a single row that fails validation; carries the field errors."""

    def __init__(self, errors):
        self.errors = errors
        super().__init__(str(errors))


@dataclass
class RowError:
    row: int
    errors: dict


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list = field(default_factory=list)

    def as_dict(self):
        return {
            'created': self.created,
            'updated': self.updated,
            'errors': [{'row': row_error.row, 'errors': row_error.errors} for row_error in self.errors],
        }


class ItemExcelImporter:
    """Upserts Item rows parsed from an .xlsx file into a single company."""

    def __init__(self, company_id):
        self.company_id = company_id

    def import_workbook(self, file_obj):
        try:
            workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
            sheet = workbook[ITEMS_SHEET_NAME] if ITEMS_SHEET_NAME in workbook.sheetnames else workbook.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
        except Exception as exc:
            raise InvalidWorkbookError(f'No se pudo leer el archivo: {exc}') from exc

        result = ImportResult()
        for row_number, raw_row in enumerate(rows, start=2):
            if raw_row is None or all(cell in (None, '') for cell in raw_row):
                continue
            row_data = dict(zip(COLUMN_ORDER, raw_row))
            try:
                with transaction.atomic():
                    created = self._upsert_row(row_data)
                result.created += int(created)
                result.updated += int(not created)
            except RowValidationError as exc:
                result.errors.append(RowError(row=row_number, errors=exc.errors))

        return result

    def _upsert_row(self, row_data):
        errors = {}

        base_code = self._clean_str(row_data.get(BASE_CODE))
        base_name = self._clean_str(row_data.get(BASE_NAME))
        name = self._clean_str(row_data.get(NAME))
        if not base_code:
            errors[BASE_CODE] = REQUIRED_FIELD_LABEL
        if not base_name:
            errors[BASE_NAME] = REQUIRED_FIELD_LABEL
        if not name:
            errors[NAME] = REQUIRED_FIELD_LABEL

        item_group = self._resolve_by_name(ItemGroup, ITEM_GROUP, row_data.get(ITEM_GROUP), errors, required=True)
        unit_measure = self._resolve_unit_measure(row_data.get(UNIT_MEASURE), errors)
        iva_type, iva_rate = self._resolve_tax_pair(
            IvaType, IvaRate, 'iva_type', IVA_TYPE, IVA_RATE,
            row_data.get(IVA_TYPE), row_data.get(IVA_RATE), errors,
        )
        excise_tax_type, excise_tax_rate = self._resolve_tax_pair(
            ExciseTaxType, ExciseTaxRate, 'excise_tax_type', EXCISE_TAX_TYPE, EXCISE_TAX_RATE,
            row_data.get(EXCISE_TAX_TYPE), row_data.get(EXCISE_TAX_RATE), errors,
        )

        if errors:
            raise RowValidationError(errors)

        _, created = Item.objects.update_or_create(
            company_id=self.company_id,
            code=base_code,
            defaults={
                'base_code': base_code,
                'base_name': base_name,
                'name': name,
                'item_group': item_group,
                'unit_measure': unit_measure,
                'iva_type': iva_type,
                'iva_rate': iva_rate,
                'excise_tax_type': excise_tax_type,
                'excise_tax_rate': excise_tax_rate,
            },
        )
        return created

    @staticmethod
    def _clean_str(value):
        if value is None:
            return ''
        return str(value).strip()

    @staticmethod
    def _parse_decimal(value):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return None

    def _resolve_by_name(self, model, field_name, raw_value, errors, *, required, lookup_field='name'):
        value = self._clean_str(raw_value)
        if not value:
            if required:
                errors[field_name] = REQUIRED_FIELD_LABEL
            return None
        try:
            return model.objects.get(**{f'{lookup_field}__iexact': value})
        except model.DoesNotExist:
            errors[field_name] = f'"{value}" no encontrado.'
        except model.MultipleObjectsReturned:
            errors[field_name] = f'"{value}" es ambiguo, hay varios registros con ese nombre.'
        return None

    def _resolve_unit_measure(self, raw_value, errors):
        return self._resolve_by_name(
            UnitMeasure, UNIT_MEASURE, raw_value, errors, required=True, lookup_field='abbreviation',
        )

    def _resolve_tax_pair(self, type_model, rate_model, rate_fk_name, type_field, rate_field, raw_type, raw_rate, errors):
        type_value = self._clean_str(raw_type)
        has_rate = raw_rate not in (None, '')

        if not type_value:
            if has_rate:
                errors[type_field] = 'Se requiere el tipo para resolver la tarifa.'
            return None, None

        type_obj = self._resolve_by_name(type_model, type_field, type_value, errors, required=False)
        if type_obj is None or not has_rate:
            return type_obj, None

        rate_value = self._parse_decimal(raw_rate)
        if rate_value is None:
            errors[rate_field] = f'"{raw_rate}" no es un número válido.'
            return type_obj, None

        try:
            rate_obj = rate_model.objects.get(**{rate_fk_name: type_obj}, value=rate_value)
        except rate_model.DoesNotExist:
            errors[rate_field] = f'Tarifa "{rate_value}" no encontrada para "{type_value}".'
            return type_obj, None
        except rate_model.MultipleObjectsReturned:
            errors[rate_field] = f'Tarifa "{rate_value}" es ambigua para "{type_value}".'
            return type_obj, None

        return type_obj, rate_obj
