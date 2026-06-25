"""Renders ItemPrice querysets as .xlsx workbooks for download."""

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from .item_price_excel_columns import HEADER_ROW, ITEM_PRICES_SHEET_NAME, PRICE_TYPE_LABELS

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class ItemPriceExcelExporter:
    """Builds an .xlsx workbook from a queryset of ItemPrice records."""

    def __init__(self, queryset):
        self.queryset = queryset.select_related('item')

    def build_workbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = ITEM_PRICES_SHEET_NAME
        sheet.append(HEADER_ROW)

        for item_price in self.queryset:
            sheet.append(self._row_for(item_price))

        for index, header in enumerate(HEADER_ROW, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(len(header) + 2, 18)

        return workbook

    @staticmethod
    def _row_for(item_price):
        return [
            item_price.id,
            item_price.item_id,
            item_price.item.name if item_price.item_id else None,
            PRICE_TYPE_LABELS.get(item_price.item_price_type, item_price.item_price_type),
            float(item_price.price) if item_price.price is not None else None,
            float(item_price.iva) if item_price.iva is not None else None,
            float(item_price.excise_tax) if item_price.excise_tax is not None else None,
            float(item_price.total) if item_price.total is not None else None,
        ]

    def as_http_response(self, filename='item_prices.xlsx'):
        workbook = self.build_workbook()
        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
