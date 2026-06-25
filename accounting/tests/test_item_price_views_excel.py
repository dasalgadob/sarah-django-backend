import io
from decimal import Decimal

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from accounting.models import ItemPrice
from accounting.services.item_price_excel_columns import HEADER_ROW


def _workbook_upload_file(rows, filename='precios.xlsx'):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(HEADER_ROW)
    for row in rows:
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        filename, buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@pytest.mark.django_db
class TestItemPriceDownloadEndpoint:
    def test_download_returns_xlsx_for_company_prices(self, api_client, company, item):
        item_price = ItemPrice.objects.create(
            item=item, company=company, item_price_type=ItemPrice.SELLING,
            price='100000.00', iva='19000.00', excise_tax='0.00', total='119000.00',
        )

        response = api_client.get(f'/api/companies/{company.id}/item-prices/download/')

        assert response.status_code == 200
        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == HEADER_ROW
        assert sheet[2][0].value == item_price.id
        assert sheet[2][1].value == item.id
        assert sheet[2][3].value == 'Venta'

    def test_download_excludes_other_companies_prices(self, api_client, company, other_company, item, item_group, unit_measure):
        from accounting.models import Item

        other_item = Item.objects.create(
            base_code='OTHER001', base_name='Other', name='Other',
            item_group=item_group, unit_measure=unit_measure, company=other_company,
        )
        ItemPrice.objects.create(
            item=other_item, company=other_company, item_price_type=ItemPrice.SELLING,
            price='1.00', iva='0.19', excise_tax='0.00', total='1.19',
        )

        response = api_client.get(f'/api/companies/{company.id}/item-prices/download/')

        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        assert workbook.active.max_row == 1

    def test_download_filename_includes_company_name_and_today(self, api_client, company):
        from django.utils import timezone

        response = api_client.get(f'/api/companies/{company.id}/item-prices/download/')

        today = timezone.localdate().isoformat()
        assert response['Content-Disposition'] == (
            f'attachment; filename="Precios#{company.legal_name}#{today}.xlsx"'
        )

    def test_download_requires_authentication(self, company):
        from rest_framework.test import APIClient

        response = APIClient().get(f'/api/companies/{company.id}/item-prices/download/')

        assert response.status_code == 401


@pytest.mark.django_db
class TestItemPriceUploadEndpoint:
    def test_upload_creates_price_with_computed_taxes(self, api_client, company, item, iva_rate):
        file_obj = _workbook_upload_file([
            [None, item.id, item.name, 'Venta', 100000, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data == {'created': 1, 'updated': 0, 'errors': []}
        price = ItemPrice.objects.get(company=company, item=item)
        assert price.price == Decimal('100000.00')
        assert price.iva == Decimal('19000.00')
        assert price.total == Decimal('119000.00')

    def test_upload_updates_existing_price_by_id(self, api_client, company, item, iva_rate):
        existing = ItemPrice.objects.create(
            item=item, company=company, item_price_type=ItemPrice.BUYING,
            price='1.00', iva='0.19', excise_tax='0.00', total='1.19',
        )
        file_obj = _workbook_upload_file([
            [existing.id, item.id, item.name, 'Venta', 200000, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data == {'created': 0, 'updated': 1, 'errors': []}
        existing.refresh_from_db()
        assert existing.price == Decimal('200000.00')

    def test_upload_rejects_item_from_another_company(self, api_client, company, other_company, item_group, unit_measure):
        from accounting.models import Item

        other_item = Item.objects.create(
            base_code='OTHER001', base_name='Other', name='Other',
            item_group=item_group, unit_measure=unit_measure, company=other_company,
        )
        file_obj = _workbook_upload_file([
            [None, other_item.id, other_item.name, 'Venta', 100000, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data['created'] == 0
        assert response.data['errors'][0]['errors']['item_id'] == (
            f'El producto con id {other_item.id} no existe o no pertenece a esta empresa.'
        )

    def test_upload_reports_row_errors_without_failing_whole_request(self, api_client, company, item, iva_rate):
        file_obj = _workbook_upload_file([
            [None, item.id, item.name, 'Venta', 100000, None, None, None],
            [None, None, None, 'Venta', 100000, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data['created'] == 1
        assert response.data['errors'] == [
            {'row': 3, 'errors': {'item_id': 'Este campo es obligatorio.'}},
        ]

    def test_upload_without_file_returns_400(self, api_client, company):
        response = api_client.post(f'/api/companies/{company.id}/item-prices/upload/', {}, format='multipart')

        assert response.status_code == 400
        assert response.data == {'errors': {'file': ['Este campo es obligatorio.']}}

    def test_upload_with_invalid_file_returns_400(self, api_client, company):
        bad_file = SimpleUploadedFile('precios.xlsx', b'not a real xlsx', content_type='application/octet-stream')

        response = api_client.post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': bad_file}, format='multipart',
        )

        assert response.status_code == 400
        assert 'file' in response.data['errors']

    def test_upload_requires_authentication(self, company, item):
        from rest_framework.test import APIClient

        file_obj = _workbook_upload_file([
            [None, item.id, item.name, 'Venta', 100000, None, None, None],
        ])

        response = APIClient().post(
            f'/api/companies/{company.id}/item-prices/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 401
