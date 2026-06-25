import io

import openpyxl
import pytest

from accounting.models import Item
from accounting.services.item_excel_columns import HEADER_ROW


def _workbook_upload_file(rows, filename='items.xlsx'):
    from django.core.files.uploadedfile import SimpleUploadedFile

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(HEADER_ROW)
    for row in rows:
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        filename, buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@pytest.mark.django_db
class TestItemDownloadEndpoint:
    def test_download_returns_xlsx_for_company_items(self, api_client, company, item):
        response = api_client.get(f'/api/companies/{company.id}/items/download/')

        assert response.status_code == 200
        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == HEADER_ROW
        assert sheet[2][0].value == item.base_code

    def test_download_filename_includes_company_name_and_today(self, api_client, company, item):
        from django.utils import timezone

        response = api_client.get(f'/api/companies/{company.id}/items/download/')

        today = timezone.localdate().isoformat()
        assert response['Content-Disposition'] == (
            f'attachment; filename="Productos#{company.legal_name}#{today}.xlsx"'
        )

    def test_download_exposes_content_disposition_header_via_cors(self, api_client, company, item):
        response = api_client.get(
            f'/api/companies/{company.id}/items/download/',
            HTTP_ORIGIN='http://localhost:3000',
        )

        assert 'Content-Disposition' in response['Access-Control-Expose-Headers']

    def test_download_excludes_other_companies_items(self, api_client, company, other_company, item_group, unit_measure):
        Item.objects.create(
            base_code='OTHERCO001', base_name='Other', name='Other',
            item_group=item_group, unit_measure=unit_measure, company=other_company,
        )

        response = api_client.get(f'/api/companies/{company.id}/items/download/')

        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet.max_row == 1  # header only, no items for this company

    def test_download_respects_name_filter(self, api_client, company, item_group, unit_measure):
        Item.objects.create(
            base_code='A001', base_name='Match Me', name='Match Me',
            item_group=item_group, unit_measure=unit_measure, company=company,
        )
        Item.objects.create(
            base_code='A002', base_name='Skip Me', name='Skip Me',
            item_group=item_group, unit_measure=unit_measure, company=company,
        )

        response = api_client.get(f'/api/companies/{company.id}/items/download/?name=Match')

        workbook = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet.max_row == 2
        assert sheet[2][2].value == 'Match Me'

    def test_download_requires_authentication(self, company, item):
        from rest_framework.test import APIClient

        response = APIClient().get(f'/api/companies/{company.id}/items/download/')

        assert response.status_code == 401


@pytest.mark.django_db
class TestItemUploadEndpoint:
    def test_upload_creates_item(self, api_client, company, item_group, unit_measure):
        file_obj = _workbook_upload_file([
            ['TSHIRT001', 'Cotton T-Shirt', 'Cotton T-Shirt', item_group.name, unit_measure.abbreviation, None, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/items/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data == {'created': 1, 'updated': 0, 'errors': []}
        assert Item.objects.filter(company=company, code='TSHIRT001').exists()

    def test_upload_updates_existing_item(self, api_client, company, item, item_group, unit_measure):
        file_obj = _workbook_upload_file([
            ['TSHIRT001', 'Renamed', 'Renamed', item_group.name, unit_measure.abbreviation, None, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/items/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data == {'created': 0, 'updated': 1, 'errors': []}
        item.refresh_from_db()
        assert item.base_name == 'Renamed'

    def test_upload_reports_row_errors_without_failing_whole_request(self, api_client, company, item_group, unit_measure):
        file_obj = _workbook_upload_file([
            ['VALID001', 'Valid Item', 'Valid Item', item_group.name, unit_measure.abbreviation, None, None, None, None],
            [None, 'Missing Code', 'Missing Code', item_group.name, unit_measure.abbreviation, None, None, None, None],
        ])

        response = api_client.post(
            f'/api/companies/{company.id}/items/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 200
        assert response.data['created'] == 1
        assert response.data['errors'] == [
            {'row': 3, 'errors': {'base_code': 'Este campo es obligatorio.'}},
        ]
        assert Item.objects.filter(company=company, code='VALID001').exists()

    def test_upload_without_file_returns_400(self, api_client, company):
        response = api_client.post(f'/api/companies/{company.id}/items/upload/', {}, format='multipart')

        assert response.status_code == 400
        assert response.data == {'errors': {'file': ['Este campo es obligatorio.']}}

    def test_upload_with_invalid_file_returns_400(self, api_client, company):
        from django.core.files.uploadedfile import SimpleUploadedFile

        bad_file = SimpleUploadedFile('items.xlsx', b'not a real xlsx', content_type='application/octet-stream')

        response = api_client.post(
            f'/api/companies/{company.id}/items/upload/', {'file': bad_file}, format='multipart',
        )

        assert response.status_code == 400
        assert 'file' in response.data['errors']

    def test_upload_requires_authentication(self, company, item_group, unit_measure):
        from rest_framework.test import APIClient

        file_obj = _workbook_upload_file([
            ['TSHIRT001', 'Cotton T-Shirt', 'Cotton T-Shirt', item_group.name, unit_measure.abbreviation, None, None, None, None],
        ])

        response = APIClient().post(
            f'/api/companies/{company.id}/items/upload/', {'file': file_obj}, format='multipart',
        )

        assert response.status_code == 401
